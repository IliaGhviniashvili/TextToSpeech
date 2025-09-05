import asyncio
import re
from pathlib import Path

from openpyxl import load_workbook

# Reuse your TTS logic and 101/102 voice mapping
from main import AudioSplitter


EXCEL_PATH = Path("/Users/ilia/Desktop/textToSpeech/content/MED6.xlsx")
TARGET_SHEET_NAME = "words"  # case-insensitive match
OUTPUT_DIR = Path("/Users/ilia/Downloads/medicine/audios/georgian/words")


def _norm(s):
    return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())


def _find_sheet_by_name_case_insensitive(wb, wanted):
    wn = _norm(wanted)
    for name in wb.sheetnames:
        if _norm(name) == wn:
            return wb[name]
    return wb[wb.sheetnames[0]]


def _find_or_create_audio_col(ws, header_row=1):
    headers = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = _norm(cell.value)
        if key:
            headers[key] = idx

    words_col = headers.get("words") or headers.get("word")
    dubbers_col = headers.get("dubbers") or headers.get("dubber")
    audio_col = (
        headers.get("audiofilename")
        or headers.get("audofilename")
        or headers.get("audio_file_name")
        or headers.get("audiofile")
    )

    if audio_col is None:
        audio_col = ws.max_column + 1
        ws.cell(row=header_row, column=audio_col, value="audioFileName")

    return words_col, dubbers_col, audio_col


async def synthesize_all():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel not found at: {EXCEL_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(EXCEL_PATH)
    ws = _find_sheet_by_name_case_insensitive(wb, TARGET_SHEET_NAME)

    header_row = 1
    words_col, dubbers_col, audio_col = _find_or_create_audio_col(ws, header_row=header_row)
    if not words_col:
        raise RuntimeError("Couldn't find 'words' column (case-insensitive).")

    splitter = AudioSplitter(output_dir=str(OUTPUT_DIR))

    counter = 1
    for r in range(header_row + 1, ws.max_row + 1):
        word_val = ws.cell(row=r, column=words_col).value
        if word_val is None or str(word_val).strip() == "":
            continue

        # Skip if already filled
        cur_audio = ws.cell(row=r, column=audio_col).value
        if cur_audio and str(cur_audio).strip():
            continue

        dubber_val = None
        if dubbers_col:
            dubber_val = ws.cell(row=r, column=dubbers_col).value

        dubber_id = None
        if dubber_val is not None:
            nums = re.findall(r"\d+", str(dubber_val))
            if nums:
                dubber_id = int(nums[0])

        # Always pass a valid Edge TTS voice; avoid falling back to invalid default
        if dubber_id == 101:
            voice_override = "ka-GE-GiorgiNeural"
        elif dubber_id == 102:
            voice_override = "ka-GE-EkaNeural"
        else:
            voice_override = "ka-GE-EkaNeural"

        # Generate with underlying TTS, then rename to MED6X###### and store name without extension
        out_path = await splitter.create_sentence_audio(str(word_val).strip(), sentence_id=counter, voice_override=voice_override)
        base_name = f"MED6X{counter:06d}"
        new_path = OUTPUT_DIR / f"{base_name}.mp3"

        try:
            if new_path.exists():
                new_path.unlink()
            Path(out_path).rename(new_path)
        except Exception:
            # Fallback: if rename fails, keep original path and store its stem
            new_path = Path(out_path)
            base_name = new_path.stem

        # Write without extension into Excel
        ws.cell(row=r, column=audio_col, value=base_name)

        counter += 1

    wb.save(EXCEL_PATH)
    print("Done. Wrote filenames to 'audioFileName' and saved workbook.")


if __name__ == "__main__":
    asyncio.run(synthesize_all())